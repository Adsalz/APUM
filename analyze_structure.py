import openpyxl
import os
import re
from datetime import datetime

# Ouvrir un fichier exemple
file_path = os.path.join(os.path.dirname(__file__), 'Desideratas papiers', 'FRANCON  T.XLSX')

wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb.active

print("=== ANALYSE DE LA STRUCTURE DU FICHIER ===\n")

# Afficher les 15 premières lignes complètes
for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
    print(f"Ligne {row_idx:2d}: {list(row)}")

print("\n=== RECHERCHE DES DATES ET CRÉNEAUX ===\n")

# Chercher les lignes qui contiennent des dates ou des mots-clés
date_pattern = re.compile(r'\d{1,2}[-/]\d{1,2}[-/]\d{2,4}')
creneau_keywords = ['QUART', 'RENFORT', '1er', '2ème', '3ème', '4ème']

for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
    row_text = ' '.join([str(cell) if cell is not None else '' for cell in row])

    # Chercher les dates
    if date_pattern.search(row_text):
        print(f"[DATE] Ligne {row_idx}: {list(row)}")

    # Chercher les créneaux
    if any(keyword.lower() in row_text.lower() for keyword in creneau_keywords):
        print(f"[CRÉNEAU] Ligne {row_idx}: {list(row)}")

print("\n=== COLONNES (premières valeurs non-null) ===\n")

# Analyser chaque colonne
max_col = sheet.max_column
for col_idx in range(1, min(max_col + 1, 15)):
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    col_values = []
    for row_idx in range(1, 31):
        cell_value = sheet.cell(row_idx, col_idx).value
        if cell_value is not None and str(cell_value).strip():
            col_values.append(f"R{row_idx}: {cell_value}")

    if col_values:
        print(f"Colonne {col_letter} ({col_idx}): {col_values[:5]}")  # Afficher les 5 premières valeurs
